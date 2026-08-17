import openpyxl
import re
import json

EXCEL_PATH = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/SCHEDULE SY 2026-2027 TW.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

from parse_all_authoritative_schedules import normalize_teacher_name

DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]

def clean_time(t_str):
    if not t_str:
        return ""
    s = str(t_str).strip()
    s = re.sub(r'[\r\n\t]+', ' ', s)
    s = re.sub(r'(\d{1,2}):(\d{2}):00', r'\1:\2', s)
    
    has_am = bool(re.search(r'(?i)a\.?m\.?', s))
    has_pm = bool(re.search(r'(?i)p\.?m\.?', s))
    
    s = re.sub(r'(?i)\s*a\.?\s*m\.?', '', s)
    s = re.sub(r'(?i)\s*p\.?\s*m\.?', '', s)
    s = re.sub(r'\s*[-–—]\s*', ' – ', s).strip()
    
    if not s:
        return ""
        
    parts = s.split(' – ')
    if len(parts) == 2:
        t1, t2 = parts[0].strip(), parts[1].strip()
        m1 = re.match(r'^0?(\d{1,2}):?(\d{2})?$', t1)
        m2 = re.match(r'^0?(\d{1,2}):?(\d{2})?$', t2)
        if m1 and m2:
            h1, min1 = int(m1.group(1)), m1.group(2) or '00'
            h2, min2 = int(m2.group(1)), m2.group(2) or '00'
            p1 = 'AM' if (7 <= h1 <= 11) else 'PM'
            p2 = 'AM' if (7 <= h2 <= 11) else 'PM'
            if has_pm:
                if h1 >= 12 or h1 < 7: p1 = 'PM'
                p2 = 'PM'
            elif has_am:
                p1 = 'AM'
                p2 = 'AM'
            return f"{h1:02d}:{min1} {p1} – {h2:02d}:{min2} {p2}"
        return f"{t1} – {t2}"
    elif len(parts) == 1:
        pt = parts[0].strip()
        m = re.match(r'^0?(\d{1,2}):?(\d{2})?$', pt)
        if m:
            h, mins = int(m.group(1)), m.group(2) or '00'
            p = 'AM' if (7 <= h <= 11) else 'PM'
            if has_pm: p = 'PM'
            elif has_am: p = 'AM'
            return f"{h:02d}:{mins} {p}"
        return pt
    return s

def clean_min(m_val):
    if not m_val:
        return ""
    if isinstance(m_val, (int, float)):
        return f"{int(m_val)} min."
    s = str(m_val).strip()
    if 'min' not in s.lower() and s.isdigit():
        return f"{s} min."
    return s

def parse_cell(cell_str):
    if not cell_str:
        return None
    s = str(cell_str).strip()
    if not s:
        return None
    s = re.sub(r'[\r\n\t]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    
    s_up = s.upper()
    is_break = any(k in s_up for k in ['GENERAL ASSEMBLY', 'RECESS', 'LUNCH', 'DEPARTURE', 'TRANSITION', 'SALAH', 'SHORT BREAK', 'HOMEROOM'])
    if is_break:
        return {
            'is_break': True,
            'label': s,
            'subject': s,
            'teacher': '',
            'extra': ''
        }
    
    m = re.match(r'^(.*?)\s*[-–—]\s*(Tchr\.|Teacher|Tr\.|Ust\.|Ustdz\.|Ustadh|Ustadha|Alim|Sir)?\s*([A-Za-z\s\.\'\`]+?)(\s*\(.*?\))?$', s, flags=re.IGNORECASE)
    if m:
        subj = m.group(1).strip()
        t_raw = (m.group(2) or '') + ' ' + m.group(3).strip()
        t_norm = normalize_teacher_name(t_raw)
        extra = m.group(4) or ''
        return {
            'is_break': False,
            'label': s,
            'subject': subj,
            'teacher': t_norm,
            'extra': extra.strip()
        }
        
    m2 = re.match(r'^(.*?)\s+(Tchr\.|Teacher|Tr\.|Ust\.|Ustdz\.|Ustadh|Ustadha|Alim|Sir)\s+([A-Za-z\s\.\'\`]+)$', s, flags=re.IGNORECASE)
    if m2:
        subj = m2.group(1).strip()
        t_raw = m2.group(2) + ' ' + m2.group(3).strip()
        t_norm = normalize_teacher_name(t_raw)
        return {
            'is_break': False,
            'label': s,
            'subject': subj,
            'teacher': t_norm,
            'extra': ''
        }
        
    return {
        'is_break': False,
        'label': s,
        'subject': s,
        'teacher': '',
        'extra': ''
    }

all_sections = []

sheet_configs = [
    'ELEM',
    'HS SCHED (NEW)',
    'SHS',
    'HS SCHED'
]

for sname in sheet_configs:
    ws = wb[sname]
    
    # 1. Discover all tables in sheet
    discovered = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                s = v.strip()
                s_up = s.upper()
                if any(k in s_up for k in ['GRADE', 'KINDER', 'SCHEDULE', 'K1', 'K2', 'FACE TO FACE', '1ST SHIFT', '2ND SHIFT']) and len(s) < 80 and not any(k in s_up for k in ['GENERAL ASSEMBLY', 'RECESS', 'LUNCH', 'DEPARTURE', 'TRANSITION', 'SALAH']):
                    for tr in [r+1, r+2]:
                        if tr <= ws.max_row:
                            r_vals = [str(ws.cell(row=tr, column=cc).value).lower() for cc in range(c, min(ws.max_column+1, c+5))]
                            if any('time' in x or 'mins' in x or 'minutes' in x or 'sunday' in x for x in r_vals):
                                discovered.append({
                                    'title': s,
                                    'title_row': r,
                                    'time_row': tr,
                                    'col': c
                                })
                                break
                                
    # 2. Determine non-overlapping end_row for each table in column band
    for t in discovered:
        next_t_row = None
        for other in discovered:
            if other != t and other['title_row'] > t['title_row'] and abs(other['col'] - t['col']) <= 5:
                if next_t_row is None or other['title_row'] < next_t_row:
                    next_t_row = other['title_row']
        t['end_row'] = next_t_row - 1 if next_t_row else min(ws.max_row, t['time_row'] + 17)

    # 3. Merged cell lookup
    merged_map = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_map[(r, c)] = (rng.min_row, rng.min_col, rng.max_row, rng.max_col)

    # 4. Parse periods strictly inside each table boundary
    for t in discovered:
        title = t['title']
        time_row = t['time_row']
        end_row = t['end_row']
        time_col = t['col']
        min_col = t['col'] + 1
        first_day_col = min_col + 1
        
        periods = []
        for pr in range(time_row + 1, end_row + 1):
            cell_t0 = ws.cell(row=pr, column=time_col).value
            cell_m0 = ws.cell(row=pr, column=min_col).value
            val_first = ws.cell(row=pr, column=first_day_col).value
            
            # Check if this row is empty
            if cell_t0 is None and cell_m0 is None and val_first is None:
                continue
                
            # Check if this row accidentally contains another table header
            if cell_t0 and isinstance(cell_t0, str) and any(k in cell_t0.upper() for k in ['GRADE', 'KINDER', 'SCHEDULE', 'K1', 'K2']):
                break
                
            t_val = clean_time(cell_t0)
            m_val = clean_min(cell_m0)
            
            span = merged_map.get((pr, first_day_col))
            is_merged_across_days = False
            if span and span[3] >= first_day_col + 4:
                is_merged_across_days = True
                
            parsed_first = parse_cell(val_first)
            
            if is_merged_across_days:
                periods.append({
                    'time': t_val,
                    'minutes': m_val,
                    'is_merged_all_days': True,
                    'is_break': parsed_first['is_break'] if parsed_first else False,
                    'label': parsed_first['label'] if parsed_first else str(val_first or ''),
                    'subject': parsed_first['subject'] if parsed_first else str(val_first or ''),
                    'teacher': parsed_first['teacher'] if parsed_first else '',
                    'extra': parsed_first.get('extra', '') if parsed_first else '',
                    'days': {}
                })
            else:
                row_days = {}
                has_content = False
                for didx, d in enumerate(DAYS):
                    c_val = ws.cell(row=pr, column=first_day_col + didx).value
                    parsed = parse_cell(c_val)
                    row_days[d] = parsed
                    if parsed:
                        has_content = True
                        
                if not has_content and parsed_first:
                    has_content = True
                    
                if has_content or t_val:
                    periods.append({
                        'time': t_val,
                        'minutes': m_val,
                        'is_merged_all_days': False,
                        'days': row_days
                    })

        if periods:
            s_up = title.upper()
            dept = "Elementary"
            if any(k in s_up for k in ['GRADE 7', 'GRADE 8', 'GRADE 9', 'GRADE 10', '7 & 8', '9 & 10']):
                dept = "Junior High School"
            elif any(k in s_up for k in ['GRADE 11', 'GRADE 12', '11 & 12', 'SHS', 'SEMESTER']):
                dept = "Senior High School"
                
            grade = "Grade"
            for g_num in range(12, 0, -1):
                if f"GRADE {g_num}" in s_up or f"G{g_num}" in s_up:
                    grade = f"Grade {g_num}"
                    break
                elif f"{g_num} &" in s_up or f"& {g_num}" in s_up:
                    grade = f"Grade {g_num}"
                    break
            if 'KINDER 1' in s_up or 'K1' in s_up:
                grade = "Kindergarten 1"
            elif 'KINDER 2' in s_up or 'K2' in s_up:
                grade = "Kindergarten 2"
                
            shift = "F2F"
            if '1ST SHIFT' in s_up or 'FIRST SHIFT' in s_up:
                shift = "ODL - 1ST SHIFT"
            elif '2ND SHIFT' in s_up or 'SECOND SHIFT' in s_up:
                shift = "ODL - 2ND SHIFT"
                
            all_sections.append({
                'id': f"sec_{len(all_sections)+1}",
                'sheet': sname,
                'section_name': title,
                'department': dept,
                'grade_level': grade,
                'shift': shift,
                'periods': periods
            })

# De-duplicate sections
unique_sections = []
seen_names = set()
for sec in all_sections:
    s_norm = sec['section_name'].strip().lower()
    if s_norm not in seen_names:
        seen_names.add(s_norm)
        unique_sections.append(sec)

with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/class_schedules_data.json', 'w') as f:
    json.dump(unique_sections, f, indent=2)

with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/class_schedules_data.js', 'w') as f:
    f.write(f"window.OFFICIAL_CLASS_SCHEDULES = {json.dumps(unique_sections, indent=2)};\n")
    f.write(f"const OFFICIAL_CLASS_SCHEDULES = window.OFFICIAL_CLASS_SCHEDULES;\n")

print(f"Successfully extracted {len(unique_sections)} completely separated unique section schedules!")

