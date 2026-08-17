import openpyxl
import json
from datetime import datetime, time, date

EXCEL_PATH = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/SCHEDULE SY 2026-2027 TW.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

raw_export = {
    "source_file": "OFFICIAL CLASS SCHEDULE.xlsx",
    "generated_at": datetime.now().isoformat(),
    "note": "Read-only extraction from XLSX OOXML. Cell addresses and displayed stored values are preserved.",
    "sheets": []
}

def clean_val(v):
    if isinstance(v, time):
        return v.strftime('%H:%M:%S')
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v

for sname in wb.sheetnames:
    ws = wb[sname]
    merged_list = [str(r) for r in ws.merged_cells.ranges]
    
    cells_list = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                col_letter = openpyxl.utils.get_column_letter(c)
                cell_coord = f"{col_letter}{r}"
                cells_list.append({
                    "cell": cell_coord,
                    "row": r,
                    "col": c,
                    "value": clean_val(val)
                })
                
    raw_export["sheets"].append({
        "name": sname,
        "merged_ranges": merged_list,
        "nonempty_cells": cells_list
    })

with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/OFFICIAL_CLASS_SCHEDULE_raw.json', 'w') as f:
    json.dump(raw_export, f, indent=2)

print(f"Exported OFFICIAL_CLASS_SCHEDULE_raw.json with {len(raw_export['sheets'])} sheets successfully!")
