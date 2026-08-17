import openpyxl
import re
import json

EXCEL_PATH = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/SCHEDULE SY 2026-2027 TW.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def normalize_teacher_name(raw):
    if not raw:
        return ''
    s = str(raw).strip()
    # Normalize common abbreviations
    s = re.sub(r'[\r\n\t]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    
    # Strip prefixes for normalization matching
    s_clean = re.sub(r'^(Tchr\.|Teacher|Tr\.|Ust\.|Ustdz\.|Ustadh|Ustadha|Alim|Sir)\s*', '', s, flags=re.IGNORECASE).strip()
    s_low = s_clean.lower()
    
    # Explicit Aliases
    if ('muh' in s_low and 'ali' in s_low) or s_low == 'ali' or s_low == 'muhammad ali' or s_low == 'mohammad ali':
        return 'Ustadh Ali'
    if 'jairah' in s_low or 'jayra' in s_low:
        return 'Teacher Jairah'
    if 'silfa' in s_low or 'silfah' in s_low:
        return 'Ustadha Silfah'
    if 'norhydie' in s_low or 'norhidi' in s_low:
        return 'Teacher Norhydie'
    if 'monisa' in s_low:
        return 'Teacher Monisa'
    if 'sitti' in s_low:
        return 'Teacher Sitti'
    if 'marham' in s_low:
        return 'Teacher Marham'
    if 'faidh' in s_low or 'faid' in s_low:
        return 'Ustadh Faidh'
    if 'obaydah' in s_low:
        return 'Ustadh Obaydah'
    if 'abdiraheem' in s_low or 'abdulraheem' in s_low:
        return 'Ustadh Abdiraheem'
    if 'saliha' in s_low:
        return 'Ustadha Saliha'
    if 'bustamante' in s_low:
        return 'Alim Bustamante'
    if 'mamonas' in s_low:
        return 'Alim Mamonas'
    if 'samsuddin' in s_low:
        return 'Alim Samsuddin'
    if 'abdulwahab' in s_low or 'abdul wahab' in s_low or 'abdul-wahab' in s_low:
        return 'Alim Abdulwahab'
    if 'dipatuan' in s_low:
        return 'Alim Dipatuan'
    if 'jaisam' in s_low:
        return 'Ustadh Jaisam'
    if 'raffy' in s_low:
        return 'Ustadh Raffy'
    if 'arvin' in s_low:
        return 'Teacher Arvin'
    if 'saimonah' in s_low:
        return 'Teacher Saimonah'
    if 'jenny' in s_low:
        return 'Teacher Jenny'
    if 'halnaisa' in s_low:
        return 'Teacher Halnaisa'
    if 'shanen' in s_low:
        return 'Teacher Shanen'
    if 'shirehan' in s_low or s_low == 'shi':
        return 'Teacher Shirehan'
    if 'abegail' in s_low:
        return 'Teacher Abegail'
    if 'rowena' in s_low:
        return 'Teacher Rowena'
    if 'nof' in s_low:
        return 'Teacher Nof'
    if 'thea' in s_low:
        return 'Teacher Thea'
    if 'nadzra' in s_low:
        return 'Teacher Nadzra'
    if 'sophia' in s_low:
        return 'Teacher Sophia'
    if 'ethel' in s_low:
        return 'Teacher Ethel'
    if 'mohaymen' in s_low:
        return 'Sir Mohaymen'
    if 'marie' in s_low:
        return 'Teacher Marie'
    if 'ahmad' in s_low:
        return 'Teacher Ahmad'
    if 'jerlyn' in s_low:
        return 'Teacher Jerlyn'
    if 'wendy' in s_low:
        return 'Teacher Wendy'
    if 'kat' in s_low:
        return 'Teacher Kat'
    if 'junaisa' in s_low:
        return 'Teacher Junaisa'
    if 'zara' in s_low:
        return 'Teacher Zara'
    if 'ersahad' in s_low:
        return 'Ustadh Ersahad'
    if 'hainur' in s_low:
        return 'Ustadh Hainur'
    if 'abdul karim' in s_low or 'abdulkarim' in s_low:
        return 'Alim Abdul Karim'
    if 'zuhora' in s_low:
        return 'Teacher Zuhora'
    if 'nashra' in s_low:
        return 'Teacher Nashra'
    if 'fahima' in s_low:
        return 'Teacher Fahima'
    if 'hamida' in s_low:
        return 'Teacher Hamida'
    if 'raihan' in s_low:
        return 'Teacher Raihan'
    if 'amerah' in s_low:
        return 'Teacher Amerah'
    
    return s

def extract_section_tables(ws):
    tables = []
    # Identify tables by looking for 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday'
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip().lower() == 'sunday':
                # Candidate table
                # Days are cols: c (Sunday), c+1 (Monday), c+2 (Tuesday), c+3 (Wednesday), c+4 (Thursday)
                # Time is usually c-2 or c-1
                time_col = c - 2 if c >= 3 and ws.cell(row=r, column=c-2).value and 'time' in str(ws.cell(row=r, column=c-2).value).lower() else (c - 1 if c >= 2 else None)
                min_col = c - 1 if time_col == c - 2 else None
                
                # Search upward for section or teacher title
                title = None
                for pr in range(max(1, r - 4), r):
                    for pc in range(max(1, c - 2), min(ws.max_column + 1, c + 6)):
                        pv = ws.cell(row=pr, column=pc).value
                        if pv and str(pv).strip():
                            pvs = str(pv).strip()
                            if not any(k in pvs.lower() for k in ['time', 'minutes', 'sunday', 'monday', 'general assembly']):
                                title = pvs
                                break
                    if title:
                        break
                
                # Scan table rows downwards until empty time or next table
                periods = []
                for tr in range(r + 1, ws.max_row + 1):
                    time_val = ws.cell(row=tr, column=time_col).value if time_col else None
                    day_vals = [ws.cell(row=tr, column=dc).value for dc in range(c, c + 5)]
                    
                    # Stop if entirely empty
                    if not time_val and not any(day_vals):
                        break
                    # Stop if new header row
                    if time_val and 'time' in str(time_val).lower():
                        break
                    if any(str(dv).strip().lower() == 'sunday' for dv in day_vals if dv):
                        break
                        
                    min_val = ws.cell(row=tr, column=min_col).value if min_col else None
                    
                    periods.append({
                        'row': tr,
                        'time': str(time_val).strip() if time_val else '',
                        'minutes': min_val,
                        'days': {
                            'Sunday': str(day_vals[0]).strip() if day_vals[0] else '',
                            'Monday': str(day_vals[1]).strip() if day_vals[1] else '',
                            'Tuesday': str(day_vals[2]).strip() if day_vals[2] else '',
                            'Wednesday': str(day_vals[3]).strip() if day_vals[3] else '',
                            'Thursday': str(day_vals[4]).strip() if day_vals[4] else ''
                        }
                    })
                    
                tables.append({
                    'sheet': ws.title,
                    'header_row': r,
                    'start_col': c,
                    'title': title or f"Table_R{r}_C{c}",
                    'time_col': time_col,
                    'periods': periods
                })
    return tables

all_tables = []
for sname in wb.sheetnames:
    all_tables.extend(extract_section_tables(wb[sname]))

print(f"Extracted {len(all_tables)} total schedule tables across all sheets.")

with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/extracted_raw_tables.json', 'w') as f:
    json.dump(all_tables, f, indent=2, default=str)

