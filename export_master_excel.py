import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/exam_data.json', 'r') as f:
    records = json.load(f)

df = pd.DataFrame(records)

excel_path = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/Term_Examination_Schedule_S.Y._2026-2027_Optimized.xlsx'
wb = Workbook()

# Colors
NAVY = "1E293B"
BLUE = "2563EB"
EMERALD = "059669"
LIGHT_BG = "F8FAFC"
WHITE = "FFFFFF"
GRAY_TEXT = "64748B"
BORDER_COLOR = "CBD5E1"

thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)

# -------------------------------------------------------------
# Sheet 1: Master Exam Table
# -------------------------------------------------------------
ws1 = wb.active
ws1.title = "Master Exam Schedule"
ws1.views.sheetView[0].showGridLines = True

# Title block
ws1.merge_cells('A1:J1')
ws1['A1'] = "AL MUNAWWARA ISLAMIC SCHOOL"
ws1['A1'].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
ws1['A1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

ws1.merge_cells('A2:J2')
ws1['A2'] = "TERM EXAMINATION SCHEDULE — S.Y. 2026–2027"
ws1['A2'].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
ws1['A2'].fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 24

ws1.merge_cells('A3:J3')
ws1['A3'] = "Official Exam Week: September 2–3 & September 9–10, 2026 • 60 Minutes per Subject • Continuous Flow"
ws1['A3'].font = Font(name="Arial", size=10, italic=True, color="334155")
ws1['A3'].fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
ws1['A3'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[3].height = 20

headers = ["Date", "Day", "Exam Day", "Time Window", "Duration", "Grade Level", "Section", "Modality & Shift", "Subject", "Assigned Teacher / Proctor"]
for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=5, column=col_idx, value=h)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border
ws1.row_dimensions[5].height = 26

sorted_recs = sorted(records, key=lambda x: (x['date'], x['modality'], x['grade'], x['section'], x['startTime']))

for r_idx, r in enumerate(sorted_recs, 6):
    shift_str = f"{r['modality']} — {r['shift']}" if r['modality'] == 'ODL' else "F2F (Classroom)"
    row_data = [
        r['date'],
        r['dayName'],
        r['examDay'],
        r['time'],
        r['duration'],
        r['grade'],
        r['cleanSection'],
        shift_str,
        r['subject'],
        r['teacher']
    ]
    fill_c = PatternFill(start_color=WHITE if r_idx % 2 == 0 else LIGHT_BG, end_color=WHITE if r_idx % 2 == 0 else LIGHT_BG, fill_type="solid")
    for col_idx, val in enumerate(row_data, 1):
        c = ws1.cell(row=r_idx, column=col_idx, value=val)
        c.font = Font(name="Arial", size=9)
        c.fill = fill_c
        c.border = thin_border
        if col_idx in [1, 2, 3, 4, 5]:
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[r_idx].height = 20

# -------------------------------------------------------------
# Sheet 2: Section Daily Summary
# -------------------------------------------------------------
ws2 = wb.create_sheet(title="Section Daily Flow")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells('A1:H1')
ws2['A1'] = "SECTION DAILY FLOW & DONE-FOR-THE-DAY SUMMARY"
ws2['A1'].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
ws2['A1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

s_headers = ["Grade", "Section", "Modality", "Date", "Day", "First Exam Start", "Final Exam End (DONE FOR THE DAY)", "Total Exams"]
for col_idx, h in enumerate(s_headers, 1):
    c = ws2.cell(row=3, column=col_idx, value=h)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border
ws2.row_dimensions[3].height = 24

sec_days = defaultdict(list)
for r in records:
    k = (r['grade'], r['cleanSection'], r['modality'], r['date'], r['dayName'])
    sec_days[k].append(r)

row_idx = 4
for (g, sec, mod, d_date, d_name), r_list in sorted(sec_days.items()):
    r_list.sort(key=lambda x: x['startTime'])
    first_st = r_list[0]['startTime']
    last_et = r_list[-1]['endTime']
    total_count = len(r_list)

    fill_c = PatternFill(start_color=WHITE if row_idx % 2 == 0 else LIGHT_BG, end_color=WHITE if row_idx % 2 == 0 else LIGHT_BG, fill_type="solid")
    row_vals = [g, sec, mod, d_date, d_name, first_st, f"{last_et} (DONE FOR THE DAY)", total_count]
    for c_idx, val in enumerate(row_vals, 1):
        c = ws2.cell(row=row_idx, column=c_idx, value=val)
        c.font = Font(name="Arial", size=9)
        c.fill = fill_c
        c.border = thin_border
        if c_idx in [4, 5, 6, 7, 8]:
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[row_idx].height = 20
    row_idx += 1

# -------------------------------------------------------------
# Sheet 3: Teacher Workload & Schedule
# -------------------------------------------------------------
ws3 = wb.create_sheet(title="Teacher Assignments")
ws3.views.sheetView[0].showGridLines = True

ws3.merge_cells('A1:F1')
ws3['A1'] = "FACULTY EXAM INVENTORY & LOAD DISTRIBUTION"
ws3['A1'].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

t_headers = ["Teacher Name", "Total Assigned Exams", "Day 1 (Sep 2)", "Day 2 (Sep 3)", "Day 3 (Sep 9)", "Day 4 (Sep 10)"]
for col_idx, h in enumerate(t_headers, 1):
    c = ws3.cell(row=3, column=col_idx, value=h)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=EMERALD, end_color=EMERALD, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border
ws3.row_dimensions[3].height = 24

t_map = defaultdict(lambda: defaultdict(int))
for r in records:
    t_map[r['teacher']][r['date']] += 1

row_idx = 4
for t_name, d_counts in sorted(t_map.items()):
    total_e = sum(d_counts.values())
    d1 = d_counts.get("2026-09-02", 0)
    d2 = d_counts.get("2026-09-03", 0)
    d3 = d_counts.get("2026-09-09", 0)
    d4 = d_counts.get("2026-09-10", 0)

    fill_c = PatternFill(start_color=WHITE if row_idx % 2 == 0 else LIGHT_BG, end_color=WHITE if row_idx % 2 == 0 else LIGHT_BG, fill_type="solid")
    row_vals = [t_name, total_e, d1, d2, d3, d4]
    for c_idx, val in enumerate(row_vals, 1):
        c = ws3.cell(row=row_idx, column=c_idx, value=val)
        c.font = Font(name="Arial", size=9)
        c.fill = fill_c
        c.border = thin_border
        if c_idx > 1:
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[row_idx].height = 20
    row_idx += 1

# Auto-fit column widths
for ws in [ws1, ws2, ws3]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save(excel_path)
print(f"Master Excel workbook saved to: {excel_path}")

# Also copy to Downloads folder
import shutil
downloads_xlsx = '/home/tatsuya/Downloads/Term_Examination_Schedule_S.Y._2026-2027_Optimized.xlsx'
shutil.copy(excel_path, downloads_xlsx)
print(f"Copied Excel to Downloads: {downloads_xlsx}")
