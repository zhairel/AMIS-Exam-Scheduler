import openpyxl
import re

wb = openpyxl.load_workbook('/home/tatsuya/Projects/AMIS/amis_exam_calendar/SCHEDULE SY 2026-2027 TW.xlsx', data_only=True)

for sname in ['ELEM', 'HS SCHED (NEW)', 'SHS', 'HS SCHED']:
    ws = wb[sname]
    tables = []
    
    # 1. Discover all table header origins
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                s = v.strip().upper()
                if any(k in s for k in ['GRADE', 'KINDER', 'SCHEDULE', 'K1', 'K2']) and len(s) < 80 and not any(k in s for k in ['GENERAL ASSEMBLY', 'RECESS', 'LUNCH', 'DEPARTURE']):
                    # Check if r+1 or r+2 has Time
                    for tr in [r+1, r+2]:
                        if tr <= ws.max_row:
                            r_vals = [str(ws.cell(row=tr, column=cc).value).lower() for cc in range(c, min(ws.max_column+1, c+5))]
                            if any('time' in x or 'mins' in x or 'minutes' in x or 'sunday' in x for x in r_vals):
                                tables.append({
                                    'title': v.strip(),
                                    'title_row': r,
                                    'time_row': tr,
                                    'col': c,
                                    'end_row': None
                                })
                                break

    # 2. Determine exact end_row for each table (by next table in same col range or sheet max)
    # Sort tables by col, then title_row
    tables.sort(key=lambda t: (t['col'], t['title_row']))
    
    for i, t in enumerate(tables):
        # find next table that shares the same column (or within col..col+5)
        next_t_row = None
        for other in tables:
            if other != t and other['title_row'] > t['title_row'] and abs(other['col'] - t['col']) <= 5:
                if next_t_row is None or other['title_row'] < next_t_row:
                    next_t_row = other['title_row']
        
        t['end_row'] = next_t_row - 1 if next_t_row else min(ws.max_row, t['time_row'] + 18)

    print(f"\n=== Sheet: {sname} ({len(tables)} tables) ===")
    for t in tables:
        print(f"  Col {t['col']:2d} | Rows {t['time_row']+1:3d}..{t['end_row']:3d} | {t['title']}")

