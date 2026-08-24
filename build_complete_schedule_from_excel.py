import openpyxl
import re
import json
import os

EXCEL_PATH = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/SCHEDULE SY 2026-2027 TW.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def normalize_teacher_name(raw):
    if not raw:
        return ''
    s = str(raw).strip()
    s = re.sub(r'[\r\n\t]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    
    # Strip prefix
    s_clean = re.sub(r'^(Tchr\.|Teacher|Tr\.|Ust\.|Ustdz\.|Ustadh|Ustadha|Alim|Sir)\s*', '', s, flags=re.IGNORECASE).strip()
    s_low = s_clean.lower()
    
    # Exact / Aliases matching
    if ('muh' in s_low and 'ali' in s_low) or s_low == 'ali' or s_low == 'muhammad ali' or s_low == 'mohammad ali' or 'ust ali' in s_low:
        return 'Ustadh Ali'
    if 'jairah' in s_low or 'jayra' in s_low:
        return 'Teacher Jairah'
    if 'silfa' in s_low or 'silfah' in s_low:
        return 'Ustadha Silfah'
    if 'norhydie' in s_low or 'norhidi' in s_low:
        return 'Teacher Norhydie'
    if 'monisa' in s_low:
        return 'Teacher Monisa'
    if 'sitti' in s_low:
        return 'Teacher Sitti'
    if 'marham' in s_low:
        return 'Teacher Marham'
    if 'faidh' in s_low or 'faid' in s_low:
        return 'Ustadh Faidh'
    if 'obaydah' in s_low or 'obayda' in s_low or 'ubaydah' in s_low:
        return 'Ustadh Obaydah'
    if 'abdiraheem' in s_low or 'abdulraheem' in s_low or 'abdi' in s_low:
        return 'Ustadh Abdiraheem'
    if 'saliha' in s_low:
        return 'Ustadha Saliha'
    if 'bustamante' in s_low:
        return 'Alim Bustamante'
    if 'mamonas' in s_low:
        return 'Alim Mamonas'
    if 'samsuddin' in s_low:
        return 'Alim Samsuddin'
    if 'abdulwahab' in s_low or 'abdul wahab' in s_low or 'abdul-wahab' in s_low:
        return 'Alim Abdulwahab'
    if 'dipatuan' in s_low:
        return 'Alim Dipatuan'
    if 'jaisam' in s_low or 'jaesam' in s_low:
        return 'Ustadh Jaisam'
    if 'raffy' in s_low:
        return 'Ustadh Raffy'
    if 'arvin' in s_low:
        return 'Teacher Arvin'
    if 'saimonah' in s_low or 'saimona' in s_low:
        return 'Teacher Saimonah'
    if 'jenny' in s_low:
        return 'Teacher Jenny'
    if 'halnaisa' in s_low:
        return 'Teacher Halnaisa'
    if 'shanen' in s_low:
        return 'Teacher Shanen'
    if 'shirehan' in s_low or s_low == 'shi':
        return 'Teacher Shirehan'
    if 'abegail' in s_low:
        return 'Teacher Abegail'
    if 'rowena' in s_low:
        return 'Teacher Rowena'
    if 'nof' in s_low:
        return 'Teacher Nof'
    if 'thea' in s_low:
        return 'Teacher Thea'
    if 'nadzra' in s_low:
        return 'Teacher Nadzra'
    if 'sophia' in s_low:
        return 'Teacher Sophia'
    if 'ethel' in s_low:
        return 'Teacher Ethel'
    if 'mohaymen' in s_low or 'moh' in s_low:
        return 'Sir Mohaymen'
    if 'marie' in s_low:
        return 'Teacher Marie'
    if 'ahmad' in s_low:
        return 'Teacher Ahmad'
    if 'jerlyn' in s_low:
        return 'Teacher Jerlyn'
    if 'wendy' in s_low or 'wendelyn' in s_low:
        return 'Teacher Wendy'
    if 'kat' in s_low:
        return 'Teacher Kat'
    if 'junaisa' in s_low:
        return 'Teacher Junaisa'
    if 'zara' in s_low:
        return 'Teacher Zara'
    if 'ersahad' in s_low:
        return 'Ustadh Ersahad'
    if 'hainur' in s_low:
        return 'Ustadha Hainur'
    if 'abdul karim' in s_low or 'abdulkarim' in s_low:
        return 'Alim Abdul Karim'
    if 'zuhora' in s_low:
        return 'Teacher Zuhora'
    if 'nashra' in s_low:
        return 'Teacher Nashra'
    if 'fahima' in s_low:
        return 'Teacher Fahima'
    if 'hamida' in s_low:
        return 'Teacher Hamida'
    if 'raihan' in s_low:
        return 'Teacher Raihan'
    if 'amerah' in s_low:
        return 'Teacher Amerah'
    if 'keychell' in s_low:
        return 'Teacher Keychelle'
    if 'normylah' in s_low or 'normayla' in s_low:
        return 'Teacher Normylah'
    if 'fhairudz' in s_low or 'fairudz' in s_low:
        return 'Teacher Fhairudz'
    if 'joanna' in s_low:
        return 'Teacher Joanna'
    if 'sahdia' in s_low:
        return 'Teacher Sahdia'
    if 'anna' in s_low:
        return 'Teacher Anna'
    if 'jessa' in s_low:
        return 'Teacher Jessa'
    if 'wardah' in s_low:
        return 'Teacher Wardah'

    if not s_clean or s_clean.lower() == 'teacher':
        return ''

    return f"Teacher {s_clean.title()}"

# Standard time slots for Faculty Timetable
STANDARD_TIME_SLOTS = [
    # F2F MORNING
    {"time": "07:30 - 07:40 AM", "minutes": 10, "type": "break", "label": "GENERAL ASSEMBLY (F2F)"},
    {"time": "07:40 - 08:25 AM", "minutes": 45, "type": "period", "period_num": 1, "session": "F2F"},
    {"time": "08:25 - 09:05 AM", "minutes": 40, "type": "period", "period_num": 2, "session": "F2F"},
    {"time": "09:05 - 09:45 AM", "minutes": 40, "type": "period", "period_num": 3, "session": "F2F"},
    {"time": "09:45 - 10:00 AM", "minutes": 15, "type": "break", "label": "RECESS"},
    {"time": "10:00 - 10:45 AM", "minutes": 45, "type": "period", "period_num": 4, "session": "F2F"},
    {"time": "10:45 - 11:30 AM", "minutes": 45, "type": "period", "period_num": 5, "session": "F2F"},
    {"time": "11:30 AM - 12:30 PM", "minutes": 60, "type": "break", "label": "LUNCH & DHUHR SALAH"},
    # ODL 1ST SHIFT
    {"time": "12:30 - 12:40 PM", "minutes": 10, "type": "break", "label": "GENERAL ASSEMBLY (1ST SHIFT)"},
    {"time": "12:40 - 01:25 PM", "minutes": 45, "type": "period", "period_num": 6, "session": "ODL 1ST SHIFT"},
    {"time": "01:25 - 02:10 PM", "minutes": 45, "type": "period", "period_num": 7, "session": "ODL 1ST SHIFT"},
    {"time": "02:15 - 03:00 PM", "minutes": 45, "type": "period", "period_num": 8, "session": "ODL 1ST SHIFT"},
    {"time": "03:00 - 03:30 PM", "minutes": 30, "type": "break", "label": "ASR SALAH BREAK"},
    # ODL 2ND SHIFT
    {"time": "03:30 - 03:40 PM", "minutes": 10, "type": "break", "label": "GENERAL ASSEMBLY (2ND SHIFT)"},
    {"time": "03:40 - 04:20 PM", "minutes": 40, "type": "period", "period_num": 9, "session": "ODL 2ND SHIFT"},
    {"time": "04:30 - 05:10 PM", "minutes": 40, "type": "period", "period_num": 10, "session": "ODL 2ND SHIFT"},
    {"time": "05:20 - 06:00 PM", "minutes": 40, "type": "period", "period_num": 11, "session": "ODL 2ND SHIFT"}
]

DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]

all_teacher_grid = {}

def get_or_create_teacher_grid(t_name):
    if t_name not in all_teacher_grid:
        all_teacher_grid[t_name] = {
            'teacher_name': t_name,
            'schedule': []
        }
        for st in STANDARD_TIME_SLOTS:
            all_teacher_grid[t_name]['schedule'].append({
                'time': st['time'],
                'minutes': st['minutes'],
                'type': st['type'],
                'label': st.get('label', ''),
                'Sunday': None if st['type'] == 'period' else st.get('label', ''),
                'Monday': None if st['type'] == 'period' else st.get('label', ''),
                'Tuesday': None if st['type'] == 'period' else st.get('label', ''),
                'Wednesday': None if st['type'] == 'period' else st.get('label', ''),
                'Thursday': None if st['type'] == 'period' else st.get('label', '')
            })
    return all_teacher_grid[t_name]

def match_time_slot(time_str):
    if not time_str:
        return None
    ts = str(time_str).lower()
    if '7:40' in ts: return 1
    if '8:25' in ts: return 2
    if '9:05' in ts: return 3
    if '10:00' in ts: return 5
    if '10:45' in ts: return 6
    if '12:40' in ts: return 9
    if '1:25' in ts or '1:30' in ts: return 10
    if '2:15' in ts or '2:20' in ts: return 11
    if '3:40' in ts: return 14
    if '4:30' in ts: return 15
    if '5:20' in ts: return 16
    return None

# Parse dedicated teacher tables in ISAL UPDATED
isal_sheet = wb['ISAL UPDATED']
isal_teachers = [
    (2, 2, 'USTADH ABDIRAHEEM'),
    (2, 10, 'ALIM BUSTAMANTE'),
    (23, 2, 'Ustadh Jaisam'),
    (23, 10, 'USTADH OBAYDAH'),
    (43, 2, 'USTADH ALI'),
    (43, 10, 'USTADH RAFFY')
]

for start_r, start_c, raw_name in isal_teachers:
    t_name = normalize_teacher_name(raw_name)
    t_obj = get_or_create_teacher_grid(t_name)
    # Days are cols start_c+2 .. start_c+6
    time_col = start_c
    for r in range(start_r + 2, start_r + 20):
        time_val = isal_sheet.cell(row=r, column=time_col).value
        slot_idx = match_time_slot(time_val)
        if slot_idx is not None:
            for day_idx, day in enumerate(DAYS):
                cell_val = isal_sheet.cell(row=r, column=start_c + 2 + day_idx).value
                if cell_val and str(cell_val).strip() and str(cell_val).strip() not in ['GENERAL ASSEMBLY', 'RECESS', 'LUNCH and SALAH', 'SALAH & DEPARTURE', 'Transition', 'General Assembly']:
                    t_obj['schedule'][slot_idx][day] = str(cell_val).strip()

# Parse dedicated teacher tables in HS LOADS
hs_loads_sheet = wb['HS LOADS']
hs_teachers = [
    (2, 2, 'Teacher Jayra'),
    (2, 10, 'Teacher Aniah'),
    (2, 18, 'Teacher Halnaisa'),
    (22, 2, 'Teacher Shanen'),
    (22, 10, 'Teacher Shirehan'),
    (22, 18, 'Teacher Abegail'),
    (42, 2, 'Teacher Rowena'),
    (42, 10, 'Teacher Nof'),
    (42, 18, 'Teacher Thea'),
    (62, 2, 'Teacher Nadzra'),
    (62, 10, 'Teacher Sophia'),
    (62, 18, 'Teacher Wardah'),
    (82, 2, 'Teacher Ethel'),
    (82, 10, 'Sir Mohaymen'),
    (82, 18, 'Teacher Marie')
]

for start_r, start_c, raw_name in hs_teachers:
    t_name = normalize_teacher_name(raw_name)
    t_obj = get_or_create_teacher_grid(t_name)
    time_col = start_c
    for r in range(start_r + 2, start_r + 20):
        time_val = hs_loads_sheet.cell(row=r, column=time_col).value
        slot_idx = match_time_slot(time_val)
        if slot_idx is not None:
            for day_idx, day in enumerate(DAYS):
                cell_val = hs_loads_sheet.cell(row=r, column=start_c + 2 + day_idx).value
                if cell_val and str(cell_val).strip() and str(cell_val).strip() not in ['GENERAL ASSEMBLY', 'RECESS', 'LUNCH and SALAH', 'SALAH & DEPARTURE', 'Transition', 'General Assembly']:
                    t_obj['schedule'][slot_idx][day] = str(cell_val).strip()

# Also parse all elementary and high school sections to fill remaining faculty
with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/extracted_raw_tables.json') as f:
    raw_tables = json.load(f)

for t in raw_tables:
    if t['sheet'] in ['ELEM', 'HS SCHED (NEW)', 'SHS', 'HS SCHED']:
        sec_title = t['title']
        for p in t['periods']:
            slot_idx = match_time_slot(p['time'])
            if slot_idx is not None:
                for day in DAYS:
                    cell = p['days'].get(day)
                    if cell:
                        m = re.match(r'^(.*?)\s*[-–—]\s*(Tchr\.|Teacher|Tr\.|Ust\.|Ustdz\.|Ustadh|Ustadha|Alim|Sir)?\s*([A-Za-z\s\.\'\`]+?)(\s*\(.*?\))?$', str(cell).strip(), flags=re.IGNORECASE)
                        if m:
                            subj = m.group(1).strip()
                            t_raw = (m.group(2) or '') + ' ' + m.group(3).strip()
                            t_norm = normalize_teacher_name(t_raw)
                            if t_norm and not any(k in t_norm.upper() for k in ['GENERAL ASSEMBLY', 'BREAK', 'RECESS', 'LUNCH', 'DEPARTURE']):
                                t_obj = get_or_create_teacher_grid(t_norm)
                                if not t_obj['schedule'][slot_idx][day]:
                                    clean_sec = re.sub(r'^(GRADE|KINDER|\d+)\s*[-–—]\s*', '', sec_title, flags=re.IGNORECASE).strip()
                                    t_obj['schedule'][slot_idx][day] = f"{subj} - {clean_sec}"

print(f"Total faculty weekly schedules populated: {len(all_teacher_grid)}")
for t_name in sorted(all_teacher_grid.keys()):
    active_periods = sum(1 for p in all_teacher_grid[t_name]['schedule'] if p['type'] == 'period' and any(p[d] for d in DAYS))
    print(f"  - {t_name:<30} (Active Teaching Slots: {active_periods})")

with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/teacher_weekly_schedules.json', 'w') as f:
    json.dump(all_teacher_grid, f, indent=2)

with open('/home/tatsuya/Projects/AMIS/amis_exam_calendar/teacher_weekly_schedules.js', 'w') as f:
    f.write(f"const TEACHER_WEEKLY_SCHEDULES = {json.dumps(all_teacher_grid, indent=2)};\n")

print("Saved teacher_weekly_schedules.json and teacher_weekly_schedules.js successfully!")

