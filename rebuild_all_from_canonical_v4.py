#!/usr/bin/env python3
"""
Complete Rebuilder from Canonical Dataset V4 (SCHEDULE SY 2026-2027 TW.xlsx)
Only Sources:
  - ELEM (Kinder to Grade 6)
  - HS SCHED (NEW) (Grade 7 to Grade 10 ONLY: B1:H68, K1:Q62, S1:Y75)
  - SHS (Grade 11 to Grade 12 FIRST TERM ONLY: Rows 1:46)
"""

import json, re, os, datetime
import openpyxl

EXCEL_PATH = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/SCHEDULE SY 2026-2027 TW.xlsx'
OUTPUT_V4_JSON = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/AMIS_CLASS_DATASET_CANONICAL_LATEST_V4.json'
CLASS_DATA_JSON = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/class_schedules_data.json'
CLASS_DATA_JS = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/class_schedules_data.js'
EXAM_DATA_JSON = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/exam_data.json'
EXAM_DATA_JS = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/exam_data.js'
OPTIONS_EXAM_DATA_JSON = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/options_exam_data.json'
TEACHER_WEEKLY_JSON = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/teacher_weekly_schedules.json'
TEACHER_WEEKLY_JS = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/teacher_weekly_schedules.js'
AUDIT_LOG_PATH = '/home/tatsuya/Projects/AMIS/amis_exam_calendar/canonical_v4_teacher_audit.txt'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
DAYS_OF_WEEK = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"]

SECTION_DEFS = [
    # --- ELEM (Kinder to Grade 6) ---
    {"sheet": "ELEM", "header_cell": "B3", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 5, "end_row": 12, "shift": "F2F", "grade": "Kinder 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J3", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 5, "end_row": 11, "shift": "ODL - 1ST SHIFT", "grade": "Kinder 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R3", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 5, "end_row": 11, "shift": "ODL - 1ST SHIFT", "grade": "Kinder 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z3", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 5, "end_row": 11, "shift": "ODL - 2ND SHIFT", "grade": "Kinder 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH3", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 5, "end_row": 11, "shift": "ODL - 2ND SHIFT", "grade": "Kinder 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH13", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 15, "end_row": 21, "shift": "ODL - 2ND SHIFT", "grade": "Kinder 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z13", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 15, "end_row": 21, "shift": "ODL - 2ND SHIFT", "grade": "Kinder 1", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "B14", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 16, "end_row": 21, "shift": "F2F", "grade": "Kinder 1", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J14", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 16, "end_row": 21, "shift": "ODL - 1ST SHIFT", "grade": "Kinder 1", "dept": "Elementary"},

    {"sheet": "ELEM", "header_cell": "B24", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 26, "end_row": 37, "shift": "F2F", "grade": "Grade 1", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J24", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 26, "end_row": 32, "shift": "ODL - 1ST SHIFT", "grade": "Grade 1", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R24", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 26, "end_row": 32, "shift": "ODL - 1ST SHIFT", "grade": "Grade 1", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z24", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 26, "end_row": 32, "shift": "ODL - 2ND SHIFT", "grade": "Grade 1", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH24", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 26, "end_row": 32, "shift": "ODL - 2ND SHIFT", "grade": "Grade 1", "dept": "Elementary"},

    {"sheet": "ELEM", "header_cell": "B39", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 41, "end_row": 52, "shift": "F2F", "grade": "Grade 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J39", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 41, "end_row": 47, "shift": "ODL - 1ST SHIFT", "grade": "Grade 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R39", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 41, "end_row": 47, "shift": "ODL - 1ST SHIFT", "grade": "Grade 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z39", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 41, "end_row": 47, "shift": "ODL - 2ND SHIFT", "grade": "Grade 2", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH39", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 41, "end_row": 47, "shift": "ODL - 2ND SHIFT", "grade": "Grade 2", "dept": "Elementary"},

    {"sheet": "ELEM", "header_cell": "B54", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 56, "end_row": 67, "shift": "F2F", "grade": "Grade 3", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J54", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 56, "end_row": 62, "shift": "ODL - 1ST SHIFT", "grade": "Grade 3", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R54", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 56, "end_row": 62, "shift": "ODL - 1ST SHIFT", "grade": "Grade 3", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R64", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 66, "end_row": 72, "shift": "ODL - 1ST SHIFT", "grade": "Grade 3", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z54", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 56, "end_row": 62, "shift": "ODL - 2ND SHIFT", "grade": "Grade 3", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH54", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 56, "end_row": 62, "shift": "ODL - 2ND SHIFT", "grade": "Grade 3", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH64", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 66, "end_row": 72, "shift": "ODL - 2ND SHIFT", "grade": "Grade 3", "dept": "Elementary"},

    {"sheet": "ELEM", "header_cell": "B74", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 76, "end_row": 87, "shift": "F2F", "grade": "Grade 4", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J74", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 76, "end_row": 82, "shift": "ODL - 1ST SHIFT", "grade": "Grade 4", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R74", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 76, "end_row": 82, "shift": "ODL - 1ST SHIFT", "grade": "Grade 4", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J64", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 66, "end_row": 72, "shift": "ODL - 1ST SHIFT", "grade": "Grade 4", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z74", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 76, "end_row": 82, "shift": "ODL - 2ND SHIFT", "grade": "Grade 4", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH74", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 76, "end_row": 82, "shift": "ODL - 2ND SHIFT", "grade": "Grade 4", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AP74", "time_col": 42, "min_col": 43, "day_cols": [44,45,46,47,48], "start_row": 76, "end_row": 82, "shift": "ODL - 2ND SHIFT", "grade": "Grade 4", "dept": "Elementary"},

    {"sheet": "ELEM", "header_cell": "B89", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 91, "end_row": 102, "shift": "F2F", "grade": "Grade 5", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J89", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 91, "end_row": 97, "shift": "ODL - 1ST SHIFT", "grade": "Grade 5", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R89", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 91, "end_row": 97, "shift": "ODL - 1ST SHIFT", "grade": "Grade 5", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z89", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 91, "end_row": 97, "shift": "ODL - 1ST SHIFT", "grade": "Grade 5", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH89", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 91, "end_row": 97, "shift": "ODL - 2ND SHIFT", "grade": "Grade 5", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AP89", "time_col": 42, "min_col": 43, "day_cols": [44,45,46,47,48], "start_row": 91, "end_row": 97, "shift": "ODL - 2ND SHIFT", "grade": "Grade 5", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AX89", "time_col": 50, "min_col": 51, "day_cols": [52,53,54,55,56], "start_row": 91, "end_row": 97, "shift": "ODL - 2ND SHIFT", "grade": "Grade 5", "dept": "Elementary"},

    {"sheet": "ELEM", "header_cell": "B104", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 106, "end_row": 117, "shift": "F2F", "grade": "Grade 6", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "J104", "time_col": 10, "min_col": 11, "day_cols": [12,13,14,15,16], "start_row": 106, "end_row": 112, "shift": "ODL - 1ST SHIFT", "grade": "Grade 6", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "R104", "time_col": 18, "min_col": 19, "day_cols": [20,21,22,23,24], "start_row": 106, "end_row": 112, "shift": "ODL - 1ST SHIFT", "grade": "Grade 6", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "Z104", "time_col": 26, "min_col": 27, "day_cols": [28,29,30,31,32], "start_row": 106, "end_row": 112, "shift": "ODL - 2ND SHIFT", "grade": "Grade 6", "dept": "Elementary"},
    {"sheet": "ELEM", "header_cell": "AH104", "time_col": 34, "min_col": 35, "day_cols": [36,37,38,39,40], "start_row": 106, "end_row": 112, "shift": "ODL - 2ND SHIFT", "grade": "Grade 6", "dept": "Elementary"},

    # --- HS SCHED (NEW) (Grade 7 to Grade 10 ONLY) ---
    {"sheet": "HS SCHED (NEW)", "header_cell": "B5", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 7, "end_row": 18, "shift": "F2F", "grade": "Grade 7 & 8", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "B21", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 23, "end_row": 34, "shift": "F2F", "grade": "Grade 7 & 8", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "B37", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 39, "end_row": 50, "shift": "F2F", "grade": "Grade 9 & 10", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "B53", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 55, "end_row": 66, "shift": "F2F", "grade": "Grade 9 & 10", "dept": "Junior High School"},

    {"sheet": "HS SCHED (NEW)", "header_cell": "K5", "time_col": 11, "min_col": 12, "day_cols": [13,14,15,16,17], "start_row": 7, "end_row": 13, "shift": "ODL - 1ST SHIFT", "grade": "Grade 7", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "K16", "time_col": 11, "min_col": 12, "day_cols": [13,14,15,16,17], "start_row": 18, "end_row": 24, "shift": "ODL - 1ST SHIFT", "grade": "Grade 7", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "K28", "time_col": 11, "min_col": 12, "day_cols": [13,14,15,16,17], "start_row": 30, "end_row": 36, "shift": "ODL - 1ST SHIFT", "grade": "Grade 8", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "K40", "time_col": 11, "min_col": 12, "day_cols": [13,14,15,16,17], "start_row": 42, "end_row": 48, "shift": "ODL - 1ST SHIFT", "grade": "Grade 9", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "K52", "time_col": 11, "min_col": 12, "day_cols": [13,14,15,16,17], "start_row": 54, "end_row": 60, "shift": "ODL - 1ST SHIFT", "grade": "Grade 10", "dept": "Junior High School"},

    {"sheet": "HS SCHED (NEW)", "header_cell": "S5", "time_col": 19, "min_col": 20, "day_cols": [21,22,23,24,25], "start_row": 7, "end_row": 13, "shift": "ODL - 2ND SHIFT", "grade": "Grade 7", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "S16", "time_col": 19, "min_col": 20, "day_cols": [21,22,23,24,25], "start_row": 18, "end_row": 24, "shift": "ODL - 2ND SHIFT", "grade": "Grade 8", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "S28", "time_col": 19, "min_col": 20, "day_cols": [21,22,23,24,25], "start_row": 30, "end_row": 36, "shift": "ODL - 2ND SHIFT", "grade": "Grade 8", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "S40", "time_col": 19, "min_col": 20, "day_cols": [21,22,23,24,25], "start_row": 42, "end_row": 48, "shift": "ODL - 2ND SHIFT", "grade": "Grade 9", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "S51", "time_col": 19, "min_col": 20, "day_cols": [21,22,23,24,25], "start_row": 53, "end_row": 59, "shift": "ODL - 2ND SHIFT", "grade": "Grade 9", "dept": "Junior High School"},
    {"sheet": "HS SCHED (NEW)", "header_cell": "S62", "time_col": 19, "min_col": 20, "day_cols": [21,22,23,24,25], "start_row": 64, "end_row": 70, "shift": "ODL - 2ND SHIFT", "grade": "Grade 10", "dept": "Junior High School"},

    # --- SHS (Grade 11 & Grade 12 FIRST TERM ONLY, Rows 1:46) ---
    {"sheet": "SHS", "header_cell": "B4", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 6, "end_row": 17, "shift": "F2F", "grade": "Grade 11", "dept": "Senior High School"},
    {"sheet": "SHS", "header_cell": "K4", "time_col": 11, "min_col": 12, "day_cols": [13,14,15,16,17], "start_row": 6, "end_row": 17, "shift": "F2F", "grade": "Grade 12", "dept": "Senior High School"},
    {"sheet": "SHS", "header_cell": "B20", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 22, "end_row": 30, "shift": "ODL - 1ST SHIFT", "grade": "Grade 11", "dept": "Senior High School"},
    {"sheet": "SHS", "header_cell": "K20", "time_col": 11, "min_col": 12, "day_cols": [13,14,15,16,17], "start_row": 22, "end_row": 30, "shift": "ODL - 1ST SHIFT", "grade": "Grade 12", "dept": "Senior High School"},
    {"sheet": "SHS", "header_cell": "B33", "time_col": 2, "min_col": 3, "day_cols": [4,5,6,7,8], "start_row": 35, "end_row": 42, "shift": "ODL - 2ND SHIFT", "grade": "Grade 11", "dept": "Senior High School"}
]

BREAK_KEYWORDS = [
    'GENERAL ASSEMBLY', 'RECESS', 'TRANSITION', 'LUNCH AND SALAH', 
    'SALAH & DEPARTURE', 'DEPARTURE', 'SHORT BREAK', 'BREAK', 'SALAH', 
    'HOMEROOM GUIDANCE', 'HOMEROOM', 'DISMISSAL', 'HOMEROOM GUIDANCE/ARAL PROGRAM',
    'HOMEROOM GUIDANCE/ARAL MATH', 'QUR\'AN READING (GENERAL ASSEMBLY)'
]

def is_break_text(s):
    if not s: return True
    s_upper = s.strip().upper()
    return any(b in s_upper for b in BREAK_KEYWORDS)

def split_subject_teacher(cell_str):
    if not cell_str:
        return "", ""
    s = str(cell_str).strip()
    if is_break_text(s):
        return s, ""
    
    parts = s.split(' - ')
    if len(parts) >= 2:
        subj = parts[0].strip()
        tchr = ' - '.join(parts[1:]).strip()
        return subj, tchr
    
    if '-' in s:
        p = s.split('-')
        subj = p[0].strip()
        tchr = '-'.join(p[1:]).strip()
        return subj, tchr
        
    return s, ""

def clean_teacher_name(tname):
    if not tname:
        return ""
    t = str(tname).strip()
    t = re.sub(r'^(Tchr\.?|Teacher|Ustadh|Ustadha|Ust\.?|Alim|Sir)\s*', '', t, flags=re.IGNORECASE).strip()
    return t.title()

def clean_subject_code(s):
    if not s: return ""
    s = str(s).strip()
    s = re.sub(r'([a-zA-Z\']+)(\d+)$', r'\1', s).strip()
    return s.upper()

def format_time_label(t):
    if not t: return ""
    t = str(t).strip()
    t = t.replace('a.m.', 'AM').replace('p.m.', 'PM').replace('a.m', 'AM').replace('p.m', 'PM')
    t = re.sub(r'\s*-\s*', ' – ', t)
    return t

def get_cell_value_merged(ws, row, col):
    val = ws.cell(row, col).value
    if val is not None and str(val).strip():
        return str(val).strip()
    for rng in ws.merged_cells.ranges:
        if row >= rng.min_row and row <= rng.max_row and col >= rng.min_col and col <= rng.max_col:
            top_val = ws.cell(rng.min_row, rng.min_col).value
            if top_val is not None and str(top_val).strip():
                return str(top_val).strip()
    return ''

print("Building Canonical V4 Class Schedules...")
sections_dataset = []
audit_records = []
canonical_lookup = {}

for sdef in SECTION_DEFS:
    ws = wb[sdef['sheet']]
    hdr_val = ws[sdef['header_cell']].value
    if not hdr_val:
        continue
    sec_name = str(hdr_val).strip()
    sec_id = 'sec_' + re.sub(r'[^a-zA-Z0-9]+', '_', sec_name).strip('_').lower()
    
    periods_list = []
    
    p_num = 1
    for r in range(sdef['start_row'], sdef['end_row'] + 1):
        time_val = ws.cell(r, sdef['time_col']).value
        min_val = ws.cell(r, sdef['min_col']).value
        if not time_val:
            continue
            
        time_str = format_time_label(str(time_val).strip())
        min_str = str(min_val).strip() if min_val is not None else ""
        if min_str and not min_str.endswith('min.') and not min_str.endswith('m'):
            try:
                min_str = f"{int(float(min_str))} min."
            except:
                pass
        
        # Extract 5 days with merged cell support
        raw_vals = [get_cell_value_merged(ws, r, c) for c in sdef['day_cols']]
        
        # If all 5 days are completely empty, this period does not exist for this section (skip empty template rows)
        if all(v == '' for v in raw_vals):
            continue
            
        # If only Sunday is filled and Mon-Thu are empty, copy Sunday across all 5 days (1 cell in 5 days)
        if raw_vals[0] and all(v == '' for v in raw_vals[1:]):
            raw_vals = [raw_vals[0]] * 5
            
        day_cells = {}
        row_subjects = []
        row_teachers = []
        is_break_row = False
        
        for idx, c in enumerate(sdef['day_cols']):
            dname = DAYS_OF_WEEK[idx]
            c_str = raw_vals[idx]
            col_letter = openpyxl.utils.get_column_letter(c)
            cell_ref = f"{col_letter}{r}"
            
            subj, tchr = split_subject_teacher(c_str)
            is_break_cell = is_break_text(subj)
            
            clean_t = clean_teacher_name(tchr)
            t_id = 'tchr_' + re.sub(r'[^a-zA-Z0-9]+', '_', clean_t).strip('_').lower() if clean_t else None
            
            day_cells[dname] = {
                "raw": c_str,
                "subject": subj,
                "teacher": f"Teacher {clean_t}" if clean_t else (tchr if tchr else None),
                "teacher_id": t_id,
                "is_break": is_break_cell,
                "label": subj if is_break_cell else None,
                "source_sheet": sdef['sheet'],
                "source_cell": cell_ref
            }
            
            if is_break_cell:
                is_break_row = True
            elif subj:
                row_subjects.append(subj)
                if tchr:
                    row_teachers.append(tchr)
            
            if tchr and subj and not is_break_cell:
                c_subj = clean_subject_code(subj)
                rec = {
                    "teacher": tchr,
                    "teacher_clean": clean_t,
                    "teacher_id": t_id,
                    "section": sec_name,
                    "section_id": sec_id,
                    "subject": subj,
                    "subject_code": c_subj,
                    "source_sheet": sdef['sheet'],
                    "source_cell": f"{sdef['sheet']}!{cell_ref}",
                    "day": dname,
                    "time": time_str
                }
                audit_records.append(rec)
                
                canonical_lookup[(sec_id, c_subj)] = rec
                canonical_lookup[(sec_name.upper(), c_subj)] = rec
                canonical_lookup[(sec_id, subj.upper())] = rec
                canonical_lookup[(sec_name.upper(), subj.upper())] = rec
                
        # Determine if merged across all days
        first_day_val = day_cells["Sunday"]["raw"] if "Sunday" in day_cells else ""
        all_same = bool(first_day_val) and all(day_cells[d]["raw"] == first_day_val for d in DAYS_OF_WEEK)
        
        main_subj = row_subjects[0] if row_subjects else (first_day_val if first_day_val else "BREAK / ASSEMBLY")
        main_tchr = row_teachers[0] if row_teachers else None
        clean_main_t = clean_teacher_name(main_tchr) if main_tchr else None
        
        period_obj = {
            "period_num": p_num,
            "time": time_str,
            "minutes": min_str,
            "is_merged_all_days": all_same,
            "label": main_subj if is_break_row else None,
            "subject": main_subj,
            "subject_id": 'subj_' + re.sub(r'[^a-zA-Z0-9]+', '_', main_subj).strip('_').lower(),
            "teacher": f"Teacher {clean_main_t}" if clean_main_t else main_tchr,
            "teacher_id": 'tchr_' + re.sub(r'[^a-zA-Z0-9]+', '_', clean_main_t).strip('_').lower() if clean_main_t else None,
            "is_break": is_break_row and all_same,
            "days": day_cells,
            "source_sheet": sdef['sheet'],
            "source_cell": sdef['header_cell']
        }
        periods_list.append(period_obj)
        p_num += 1
        
    if len(periods_list) == 0:
        continue
        
    sec_obj = {
        "id": sec_id,
        "section_id": sec_id,
        "section_name": sec_name,
        "shift": sdef['shift'],
        "department": sdef['dept'],
        "grade_level": sdef['grade'],
        "total_periods": len(periods_list),
        "source_sheet": sdef['sheet'],
        "source_cell": sdef['header_cell'],
        "periods": periods_list,
        "rows": periods_list
    }
    sections_dataset.append(sec_obj)

print(f"Total Canonical Sections Built: {len(sections_dataset)}")
print(f"Total Audit Records Built: {len(audit_records)}")

with open(CLASS_DATA_JSON, 'w') as f:
    json.dump(sections_dataset, f, indent=2)
with open(CLASS_DATA_JS, 'w') as f:
    f.write('const ALL_SECTIONS_DATA = ' + json.dumps(sections_dataset, indent=2) + ';\n')

# Build Teacher Weekly Schedules
teacher_dict = {}
for a in audit_records:
    tid = a['teacher_id']
    tname = a['teacher']
    clean_t = a['teacher_clean']
    
    if tid not in teacher_dict:
        teacher_dict[tid] = {
            "teacher_id": tid,
            "teacher_name": tname,
            "canonical_name": f"Teacher {clean_t}",
            "periods": [],
            "subjects": set(),
            "sections": set(),
            "total_classes": 0,
            "total_teaching_periods": 0,
            "rows": []
        }
    teacher_dict[tid]["periods"].append({
        "section": a['section'],
        "section_id": a['section_id'],
        "subject": a['subject'],
        "day": a['day'],
        "time": a['time'],
        "source_cell": a['source_cell']
    })
    teacher_dict[tid]["subjects"].add(a['subject'])
    teacher_dict[tid]["sections"].add(a['section'])

# Deduplicate teacher periods
for tid, tdata in teacher_dict.items():
    tdata["subjects"] = sorted(list(tdata["subjects"]))
    tdata["sections"] = sorted(list(tdata["sections"]))
    tdata["total_classes"] = len(tdata["periods"])
    tdata["total_teaching_periods"] = len(tdata["periods"])

with open(TEACHER_WEEKLY_JSON, 'w') as f:
    json.dump(teacher_dict, f, indent=2)
with open(TEACHER_WEEKLY_JS, 'w') as f:
    f.write('const ALL_TEACHERS_DATA = ' + json.dumps(teacher_dict, indent=2) + ';\n')

print(f"Total Verified Teachers in Weekly Timetable: {len(teacher_dict)}")

# Update Master Exam Records strictly from V4 dataset
with open(EXAM_DATA_JSON, 'r') as f:
    exam_records = json.load(f)

SUBJECT_CANONICAL_MAP = {
    'MATHEMATICS': ['MATH', 'MATH 5'],
    'GENERAL MATHEMATICS': ['GEN MATH', 'GEN. MATH', 'GENERAL MATH'],
    'ARAL MATH': ['ARAL MATH'],
    'ARABIC LANGUAGE': ['ARABIC'],
    'ARABIC': ['ARABIC'],
    'GMRC / VALUES': ['GMRC', 'VALUES ED.', 'ESP', 'VALUES'],
    'GMRC': ['GMRC'],
    'VALUES EDUCATION': ['VALUES ED.', 'ESP', 'VALUES'],
    'TECHNOLOGY AND LIVELIHOOD EDUCATION': ['TLE', 'TLE / TVL'],
    'TLE': ['TLE', 'TLE / TVL'],
    'MAPEH': ['MAPEH'],
    'PE': ['PE', 'PE 12'],
    'PHYSICAL EDUCATION': ['PE', 'PE 12'],
    'ARALING PANLIPUNAN (AP)': ['AP', 'SOC.SCI', 'SOC SCI', 'MAKABANSA'],
    'ARALING PANLIPUNAN': ['AP', 'SOC.SCI', 'SOC SCI'],
    'MAKABANSA': ['MAKABANSA'],
    'READING AND LITERACY': ['READING AND LITERACY', 'R & L', 'LANGUAGE'],
    'LANGUAGE': ['LANGUAGE'],
    'SCIENCE': ['SCIENCE', 'SCI', 'GEN SCIENCE'],
    'GENERAL SCIENCE': ['GEN SCIENCE', 'GEN. SCIENCE', 'GENERAL SCIENCE'],
    'GENERAL BIOLOGY 1': ['GEN BIO 1', 'GEN. BIO 1'],
    'GENERAL PHYSICS 1': ['GEN. PHYSICS 1', 'GEN PHYSICS 1'],
    'PRACTICAL RESEARCH 2': ['PRAC. RES. 2', 'PRAC RES 2'],
    'MEDIA AND INFORMATION LITERACY': ['MIL'],
    '21ST CENTURY LITERATURE': ['21ST LIT.', '21ST LIT', '21ST CENTURY LITERATURE'],
    'LIFE AND CAREER SKILLS': ['LCS', 'LCS 11 2ND SHIFT'],
    'PAGBASA AT PAGSUSURI NG IBA\'T IBANG TEKSTO': ['PSKP', 'PSKP 11'],
    'ENGLISH': ['ENGLISH', 'ENG', 'EC'],
    'FILIPINO': ['FILIPINO', 'FIL'],
    'QUR\'AN': ['QUR\'AN', 'QURAN'],
    'HADITH': ['HADITH'],
    'SHAF': ['SHAF'],
    'CIRCLE TIME': ['CIRCLE TIME 1', 'CIRCLE TIME 2', 'CT 1', 'CT 2', 'CIRCLE TIME', 'MEETING TIME']
}

verified_exams = 0
unverified_exams = 0

for exam in exam_records:
    sec_name = exam.get('section_name', '').strip()
    sec_id = exam.get('section_id', '')
    raw_subj = exam.get('subject', '').strip()
    subj_upper = raw_subj.upper()
    c_subj = clean_subject_code(raw_subj)
    
    # 120-minute rule: ONLY regular High School Math gets 120 minutes!
    is_hs = any(g in exam.get('section_name', '').upper() for g in ['GRADE 7', 'GRADE 8', 'GRADE 9', 'GRADE 10'])
    is_reg_math = (c_subj == 'MATH' or raw_subj.upper() == 'MATHEMATICS') and not any(k in raw_subj.upper() for k in ['ARAL', 'GENERAL', 'CALCULUS', 'STATISTICS'])
    if is_hs and is_reg_math:
        exam['duration_minutes'] = 120
    elif exam.get('duration_minutes') == 120 and not (is_hs and is_reg_math):
        exam['duration_minutes'] = 60
        
    matched = None
    aliases = SUBJECT_CANONICAL_MAP.get(subj_upper, [subj_upper, c_subj])
    
    for a in aliases:
        if (sec_id, a) in canonical_lookup:
            matched = canonical_lookup[(sec_id, a)]
            break
        if (sec_name.upper(), a) in canonical_lookup:
            matched = canonical_lookup[(sec_name.upper(), a)]
            break
            
    if matched:
        exam['teacher'] = matched['teacher']
        exam['teacher_clean'] = matched['teacher_clean']
        exam['teacher_id'] = matched['teacher_id']
        exam['teacher_status'] = "VERIFIED"
        exam['source_cell'] = matched['source_cell']
        verified_exams += 1
    else:
        exam['teacher_status'] = "TEACHER NOT VERIFIED"
        unverified_exams += 1

print(f"Exam Records: {verified_exams} Verified, {unverified_exams} Unverified")

with open(EXAM_DATA_JSON, 'w') as f:
    json.dump(exam_records, f, indent=2)
with open(EXAM_DATA_JS, 'w') as f:
    f.write('const ALL_EXAM_RECORDS = ' + json.dumps(exam_records, indent=2) + ';\n')

print("✓ Rebuild completed successfully!")
